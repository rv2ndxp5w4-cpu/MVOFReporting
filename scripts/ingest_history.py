#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BASE_FILE = ROOT / "data" / "base_assets.json"
MANUAL_FILE = ROOT / "data" / "manual_updates.json"
DEFAULT_SOURCE_DIR = Path("/Users/danielgusev/Library/CloudStorage/Dropbox/MVOF Fund audit")

STOPWORDS = {
    "mvof",
    "fund",
    "funds",
    "inc",
    "ltd",
    "lp",
    "series",
    "stock",
    "preferred",
    "private",
    "limited",
    "technologies",
    "vehicle",
    "group",
    "opportunities",
}

ALIAS_HINTS = {
    "teachmint": ["mv opportunities india limited", "mv opportunities india"],
    "coda-project-inc": ["birdly", "plato", "grammarly", "superhuman"],
    "atom-finance-inc-series-a-preferred-stock": ["toggle ai", "reflexivity"],
    "fuse-venture-capital-partners-no-2-scsp": ["market kurly", "market kurley"],
    "mighty-angel-vehicle-pte-ltd": ["mighty buildings", "mighty"],
}


def normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def tokenize(text: str) -> list[str]:
    tokens = re.findall(r"[a-z0-9]+", text.lower())
    return [t for t in tokens if len(t) > 2 and t not in STOPWORDS]


def load_json(path: Path, fallback: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return fallback
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return fallback


def ensure_manual(path: Path) -> dict[str, Any]:
    data = load_json(path, {"assets": {}, "aliases": {}})
    data.setdefault("assets", {})
    data.setdefault("aliases", {})
    return data


def infer_date(path: Path) -> str:
    name = path.name

    # YYYYMMDD
    m = re.search(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)", name)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            pass

    # DDMMYYYY
    m = re.search(r"(?<!\d)(\d{2})(\d{2})(20\d{2})(?!\d)", name)
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            pass

    # Month YYYY
    m = re.search(
        r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s*(20\d{2})\b",
        name,
        flags=re.IGNORECASE,
    )
    if m:
        mon_map = {
            "jan": 1,
            "feb": 2,
            "mar": 3,
            "apr": 4,
            "may": 5,
            "jun": 6,
            "jul": 7,
            "aug": 8,
            "sep": 9,
            "oct": 10,
            "nov": 11,
            "dec": 12,
        }
        mo = mon_map[m.group(1).lower()[:3]]
        y = int(m.group(2))
        return dt.date(y, mo, 1).isoformat()

    return dt.date.fromtimestamp(path.stat().st_mtime).isoformat()


def infer_reporting_style(name: str) -> str:
    text = name.lower()
    if re.search(r"\bq[1-4]\b|quarter|1q|2q|3q|4q", text):
        return "quarter"
    if re.search(r"\b1h\b|\bh1\b|\b2h\b|\bh2\b|half", text):
        return "half-year"
    return "full-year"


def infer_event_type(name: str) -> str:
    text = name.lower()
    if "acquisition" in text or "acquired" in text:
        return "acquisition"
    if "rename" in text or "rebrand" in text:
        return "rename"
    if "valuation" in text or "nav" in text or "financial" in text or "audit" in text:
        return "valuation"
    if "term sheet" in text or "safe" in text:
        return "transaction"
    return "analysis"


def extract_preview(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv", ".json"}:
        return path.read_text(encoding="utf-8", errors="ignore").replace("\n", " ")[:260]

    if suffix == ".docx":
        try:
            with zipfile.ZipFile(path) as zf:
                xml = zf.read("word/document.xml").decode("utf-8", "ignore")
            texts = [t.strip() for t in re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml) if t.strip()]
            return " ".join(texts)[:260]
        except Exception:
            return f"Imported source {path.name}"

    if suffix == ".pptx":
        try:
            with zipfile.ZipFile(path) as zf:
                slide_paths = [n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")]
                txt: list[str] = []
                for sp in sorted(slide_paths)[:3]:
                    xml = zf.read(sp).decode("utf-8", "ignore")
                    txt.extend([t.strip() for t in re.findall(r"<a:t>(.*?)</a:t>", xml) if t.strip()])
            return " ".join(txt)[:260]
        except Exception:
            return f"Imported source {path.name}"

    if suffix == ".xlsx":
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            parts: list[str] = []
            for s in wb.sheetnames[:2]:
                parts.append(f"Sheet:{s}")
                ws = wb[s]
                for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
                    vals = [str(v) for v in row if v not in (None, "")]
                    if vals:
                        parts.append(" ".join(vals))
            return " | ".join(parts)[:260]
        except Exception:
            return f"Imported source {path.name}"

    return f"Imported source {path.name}"


def build_alias_pool(asset: dict[str, Any]) -> list[str]:
    base: list[str] = []
    for field in ["name", "canonical_name", "underlying_asset"]:
        if asset.get(field):
            base.append(str(asset[field]))
    for alias in asset.get("aliases", []) or []:
        base.append(str(alias))
    extra = ALIAS_HINTS.get(asset["id"], [])
    base.extend(extra)

    # Add short-brand aliases from key fields to match common filename style.
    for raw in list(base):
        words = [w for w in re.findall(r"[A-Za-z0-9]+", raw) if w]
        meaningful = [w for w in words if w.lower() not in STOPWORDS and len(w) >= 4]
        if meaningful:
            base.append(meaningful[0])
            if len(meaningful) >= 2:
                base.append(" ".join(meaningful[:2]))

    uniq = []
    seen = set()
    for item in base:
        key = item.strip().lower()
        if key and key not in seen:
            seen.add(key)
            uniq.append(item)
    return uniq


def asset_match_score(filename: str, aliases: list[str]) -> int:
    f_low = filename.lower()
    f_norm = normalize(filename)
    score = 0
    for alias in aliases:
        al = alias.lower().strip()
        if not al:
            continue
        al_norm = normalize(al)
        if not al_norm:
            continue

        if al in f_low:
            score += 100
            continue

        if al_norm in f_norm:
            score += 80
            continue

        tokens = tokenize(al)
        if tokens and all(t in f_low for t in tokens[:2]):
            score += 35

        token_hits = sum(1 for t in tokens if t in f_low)
        score += token_hits * 6

    return score


def choose_matches(path: Path, assets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    scores: list[tuple[int, dict[str, Any]]] = []
    for asset in assets:
        aliases = build_alias_pool(asset)
        score = asset_match_score(path.name, aliases)
        if score > 0:
            scores.append((score, asset))

    if not scores:
        return []

    scores.sort(key=lambda x: x[0], reverse=True)
    top = scores[0][0]
    selected = [a for s, a in scores if s >= max(70, int(top * 0.78))]
    return selected[:2]


def existing_source_paths(manual: dict[str, Any]) -> set[str]:
    out: set[str] = set()
    for data in (manual.get("assets") or {}).values():
        for ev in data.get("timeline", []) or []:
            src = ev.get("source")
            if isinstance(src, str):
                out.add(src)
    return out


def ingest(source_dir: Path, base_file: Path, manual_file: Path, dry_run: bool = False) -> dict[str, Any]:
    base = load_json(base_file, {"assets": []})
    manual = ensure_manual(manual_file)

    assets = base.get("assets", [])
    seen = existing_source_paths(manual)

    files = [
        p
        for p in source_dir.rglob("*")
        if p.is_file()
        and not p.name.startswith("~$")
        and p.name != ".DS_Store"
        and p.suffix.lower() in {".pdf", ".pptx", ".docx", ".xlsx", ".txt", ".md", ".csv", ".json"}
    ]

    added = 0
    matched_files = 0
    per_asset = defaultdict(int)

    for path in sorted(files):
        source_key = str(path)
        if source_key in seen:
            continue

        matches = choose_matches(path, assets)
        if not matches:
            continue

        matched_files += 1
        preview = extract_preview(path)
        event = {
            "date": infer_date(path),
            "label": f"Imported source: {path.name}",
            "event_type": infer_event_type(path.name),
            "reporting_style": infer_reporting_style(path.name),
            "source": source_key,
            "summary": preview,
        }

        for asset in matches:
            slot = manual["assets"].setdefault(asset["id"], {"aliases": [], "timeline": []})
            duplicate = any(
                (ev.get("source") == source_key and ev.get("label") == event["label"])
                for ev in slot.get("timeline", [])
            )
            if duplicate:
                continue
            slot["timeline"].append(event)
            per_asset[asset["id"]] += 1
            added += 1

    # keep timelines sorted for each asset
    for data in manual["assets"].values():
        data["timeline"] = sorted(data.get("timeline", []), key=lambda ev: ev.get("date", ""), reverse=True)

    if not dry_run:
        manual_file.write_text(json.dumps(manual, indent=2), encoding="utf-8")

    return {
        "scanned_files": len(files),
        "matched_files": matched_files,
        "events_added": added,
        "assets_touched": len(per_asset),
        "per_asset": dict(sorted(per_asset.items(), key=lambda kv: kv[1], reverse=True)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Bulk ingest historical files into manual timeline events")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--base", type=Path, default=BASE_FILE)
    parser.add_argument("--manual", type=Path, default=MANUAL_FILE)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    report = ingest(args.source_dir, args.base, args.manual, dry_run=args.dry_run)
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
