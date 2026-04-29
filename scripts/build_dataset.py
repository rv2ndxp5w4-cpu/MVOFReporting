#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

SECTION_MAP = {
    "FUNDS": "funds",
    "COMPANIES": "companies",
    "LOANS": "loans",
    "WRITE-OFFS": "writeoffs",
    "WRITEOFFS": "writeoffs",
    "WRITE OFFS": "writeoffs",
}

DEFAULT_PATHS = {
    "xlsx": Path("/Users/danielgusev/Library/CloudStorage/Dropbox/MVOF Fund audit/MVOF Master with Dec2025 valuations.xlsx"),
    "pptx": Path("/Users/danielgusev/Library/CloudStorage/Dropbox/MVOF Fund audit/MVOF 2026 Update.pptx"),
    "overrides": Path("/Users/danielgusev/Library/Mobile Documents/com~apple~CloudDocs/MVOF Reporting/data/canonical_overrides.json"),
    "output": Path("/Users/danielgusev/Library/Mobile Documents/com~apple~CloudDocs/MVOF Reporting/data/base_assets.json"),
}

# Portfolio Report names -> Sheet2 canonical names (normalized).
PORTFOLIO_TO_SHEET2_MAP = {
    "teachminttechnologiesprivatelimited": "mvopportunitiesindialimited",
    "codaprojectinc": "birdlyplato",
    "caliiinc": "calli",
    "caliiincseriesseed1": "calli",
    "fuseventurecapitalpartnersno2scsp": "fuseventurecapitalpartners",
    "gyankaartechnologiesprivatelimited": "gyankaartechnologiesdbapagarbook",
    "mightyangelvehiclepteltd": "mightyangelvehicle",
    "rainspvseed2preferredseriesofpennainvestments": "rain",
    "tt1productsincdbasupersapiens": "ttiproductsincdbasupersapiens",
    "helbizwheelslabsinc": "wheelslabsincseriesaipreferredstock",
    "byrdtechnologies": "byrdtechnologiesinc",
    "krisptechnologiesinc": "krisp",
    "codesignalinc": "codesignal",
    "raintechnologiesinc": "rain",
    "speechifyinc": "speechify",
    "snarkaiinc": "snarkactiveloop",
    "creoatelimited": "creoate",
    "atomfinanceinc": "atomfinance",
    "houmgroupinc": "houm",
    "atlaskitchenpteltd": "atlaskitchen",
    "byukinc": "buyk",
    "productifyinc": "productfy",
    "onevcfundiilp": "onevcfundii",
    "iseedfundiiaseriesofutsavsomaniinvestmentslp": "utsaviseedfund2",
    "backendcapitaliilp": "backendcapital",
    "shrugfundiilp": "shrug11fund",
    "1984venturesiilp": "1984ventures2",
}

STABLE_ID_BY_SHEET2_NORM = {
    "3fundlp": "3-0-fund-lp",
    "shrug11fund": "shrug-fund-ii-l-p",
    "1984ventures2": "1984-ventures-ii-l-p",
    "utsaviseedfund2": "iseed-fund-ii-a-series-of-utsav-somani-investments-lp",
    "backendcapital": "backend-capital-ii-lp",
    "onevcfundii": "onevc-fund-ii-lp",
    "mightyangelvehicle": "mighty-angel-vehicle-pte-ltd",
    "mvopportunitiesindialimited": "teachmint-technologies-private-limited",
    "speechify": "speechify-inc-safe",
    "snarkactiveloop": "snark-ai-inc-series-a-3-preferred",
    "rain": "rain-technologies-inc",
    "krisp": "krisp-technologies-inc-series-a-preferred-stock",
    "gyankaartechnologiesdbapagarbook": "gyankaar-technologies-private-limited-dba-pagarbook-series-a3",
    "birdlyplato": "coda-project-inc",
    "codesignal": "codesignal-inc",
    "creoate": "creoate-limited-advance-subscription",
    "atomfinance": "atom-finance-inc-series-a-preferred-stock",
    "fuseventurecapitalpartners": "fuse-venture-capital-partners-no-2-scsp",
    "atlaskitchen": "atlas-kitchen-pte-ltd-safe",
    "calli": "calii-inc-series-seed-1-preferred-stock",
    "houm": "houm-group-inc-converted-safe",
    "curegroupgmbh": "cure-group-gmbh",
    "byrdtechnologiesinc": "byrd-technologies",
    "buyk": "byuk-inc",
    "mightybuildings": "mighty-buildings-inc-common-stock",
    "productfy": "productify-inc-series-a1-preferred-stock",
    "ttiproductsincdbasupersapiens": "tt1-products-inc-dba-supersapiens-converted-from-tt1-note",
    "wheelslabsincseriesaipreferredstock": "helbiz-wheels-labs-inc-series-a-1-1-5-pref",
}


EXTERNAL_MARKET_INFO_BY_ID = {
    "teachmint-technologies-private-limited": "Coverage gap in verified 2025Q1-2026Q1 public corpus (mapped from MV Opportunities India Limited); additional source-backed updates recommended.",
    "speechify-inc-safe": "External market signal: strong product expansion in 2026 (podcast publishing, multimodal learning, desktop rollout) following 2025 design recognition.",
    "snark-ai-inc-series-a-3-preferred": "External market signal: Activeloop remained product-led, expanding Deep Research and scientific-discovery tooling with ecosystem partnerships.",
    "rain-technologies-inc": "External market signal: Series B financing followed by payroll/HCM integration expansion (Workday, Paylocity, partner-channel growth).",
    "krisp-technologies-inc-series-a-preferred-stock": "External market signal: sustained enterprise push in contact-centre voice AI, including interpreter/accent tools and platform packaging.",
    "gyankaar-technologies-private-limited-dba-pagarbook-series-a3": "External market signal: down-valued financing round despite cashflow-positive messaging; larger A5 follow-on closed thereafter.",
    "coda-project-inc": "External market signal: Birdly/Plato surfaced primarily through a Feb 2026 seed financing mention; public corpus otherwise thin.",
    "codesignal-inc": "External market signal: high-cadence AI hiring/learning release cycle, including multilingual interview and academy initiatives.",
    "creoate-limited-advance-subscription": "Coverage gap in verified 2025Q1-2026Q1 public corpus; no confirmed dated items in the report window.",
    "atom-finance-inc-series-a-preferred-stock": "Coverage gap in verified 2025Q1-2026Q1 public corpus; known acquisition context is mainly pre-window (2024).",
    "fuse-venture-capital-partners-no-2-scsp": "External market signal (underlying Market Kurly): strategic distribution expansion with Naver and operational rollout progression.",
    "atlas-kitchen-pte-ltd-safe": "Coverage gap in verified 2025Q1-2026Q1 public corpus; no confirmed dated external items in the report window.",
    "calii-inc-series-seed-1-preferred-stock": "Coverage gap / identity ambiguity in verified corpus (mapped from Calli/Calii); additional confirmation sources recommended.",
    "houm-group-inc-converted-safe": "External market signal: narrative shifted from profitability target to reported breakeven and multifamily expansion in early 2026.",
}




@dataclass
class Sheet2Record:
    row_no: int
    type_name: str
    name: str
    description: str
    company_snapshot: str
    comments: str
    invested: float | None
    value_2023: float | None
    value_2025: float | None
    diff: float | None
    extra_comments: str


@dataclass
class PortfolioLine:
    section: str
    name: str
    previous_name: str
    geography: str
    description: str
    nominal_holding: Any
    average_cost: float | None
    book_cost: float | None
    market_price: float | None
    market_value: float | None
    pnl: float | None
    notes: str


def slugify(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower()).strip("-")
    return value or "asset"


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dt.datetime):
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def clean_company_name(value: str) -> str:
    text = re.sub(r"\s+", " ", (value or "").strip())
    text = re.sub(r"\s*\((series|pre-series|seed|class)[^)]*\)$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*[-,]\s*(common|preferred|convertible|safe|note|series|class)\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+series\s+[a-z0-9\-]+\s+(preferred|common)\s+stock$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+(preferred|common)\s+stock$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+inc\.?\s*series\b.*$", " Inc", text, flags=re.IGNORECASE)
    text = text.strip(" -,")

    fixes = {
        "MV Opportunities India Limited": "Teachmint",
        "Birdly (Plato)": "Coda (Birdly/Plato)",
        "Calli": "Calii",
        "Buyk": "Byuk",
        "Wheels Labs Inc Series A-I Preferred Stock": "Wheels Labs Inc",
        "TTI Products, Inc, DBA Supersapiens": "Supersapiens",
    }
    return fixes.get(text, text)


def canonical_match_key(name: str) -> str:
    norm = normalize_name(clean_company_name(name))
    return PORTFOLIO_TO_SHEET2_MAP.get(norm, norm)


def parse_pptx_text(pptx_path: Path) -> list[str]:
    snippets: list[str] = []
    with zipfile.ZipFile(pptx_path) as zf:
        slide_paths = sorted(
            [
                name
                for name in zf.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            ],
            key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),
        )
        for slide in slide_paths:
            raw = zf.read(slide).decode("utf-8", "ignore")
            texts = [t.strip() for t in re.findall(r"<a:t>(.*?)</a:t>", raw) if t.strip()]
            snippets.extend(texts)
    return snippets


def load_overrides(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"assets": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"assets": {}}


def build_sheet2_records(wb: openpyxl.Workbook) -> list[Sheet2Record]:
    ws = wb["Sheet2"]
    records: list[Sheet2Record] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        records.append(
            Sheet2Record(
                row_no=int(row[0]) if row[0] is not None else 0,
                type_name=str(row[1] or "").strip(),
                name=str(row[2]).strip(),
                description=str(row[3] or "").strip(),
                company_snapshot=str(row[4] or "").strip(),
                comments=str(row[5] or "").strip(),
                invested=to_float(row[6]),
                value_2023=to_float(row[7]),
                value_2025=to_float(row[8]),
                diff=to_float(row[9]),
                extra_comments=str(row[10] or "").strip(),
            )
        )
    return records


def build_portfolio_lines(wb: openpyxl.Workbook) -> list[PortfolioLine]:
    ws = wb["Portfolio Report"]
    section = "other"
    seen: set[tuple[Any, ...]] = set()
    lines: list[PortfolioLine] = []

    for row in ws.iter_rows(values_only=True):
        marker = str(row[0]).strip().upper() if row[0] else ""
        if marker in SECTION_MAP:
            section = SECTION_MAP[marker]
            continue

        security = (row[2] or "").strip() if isinstance(row[2], str) else ""
        if not security or security.upper() in {"SECURITY", "TOTAL"}:
            continue

        key = (
            security,
            to_float(row[8]),
            to_float(row[10]),
            to_float(row[12]),
        )
        if key in seen:
            continue
        seen.add(key)

        lines.append(
            PortfolioLine(
                section=section,
                name=security,
                previous_name=(row[3] or "") if isinstance(row[3], str) else "",
                geography=(row[4] or "") if isinstance(row[4], str) else "",
                description=(row[5] or "") if isinstance(row[5], str) else "",
                nominal_holding=row[6],
                average_cost=to_float(row[7]),
                book_cost=to_float(row[8]),
                market_price=to_float(row[9]),
                market_value=to_float(row[10]),
                pnl=to_float(row[12]),
                notes=(row[15] or "") if isinstance(row[15], str) else "",
            )
        )

    return lines


def infer_instrument(line_name: str, line_description: str) -> dict[str, str]:
    text = f"{line_name} {line_description}".lower()

    instrument_type = "equity"
    if "convertible" in text or "conv." in text or "loan" in text:
        instrument_type = "convertible"
    elif "safe" in text:
        instrument_type = "safe"
    elif "note" in text:
        instrument_type = "note"
    elif "common" in text:
        instrument_type = "common"
    elif "preferred" in text:
        instrument_type = "preferred"
    elif "fund" in text:
        instrument_type = "fund"

    series_match = re.search(r"(pre-series\s*[a-z0-9\-]+|series\s*[a-z0-9\-]+|seed\s*[a-z0-9\-]+)", line_name, flags=re.IGNORECASE)
    class_match = re.search(r"(class\s*[a-z0-9\-]+)", line_name, flags=re.IGNORECASE)

    return {
        "instrument_type": instrument_type,
        "series": series_match.group(1).title() if series_match else "",
        "share_class": class_match.group(1).title() if class_match else "",
    }


def parse_name_chain(previous_name: str, fallback_name: str) -> list[str]:
    if not previous_name:
        return []
    parts = [re.sub(r"\s+", " ", p).strip() for p in re.split(r"→|->", previous_name) if p.strip()]
    cleaned: list[str] = []
    for part in parts:
        part = re.sub(r"^(acquired by|renamed to|renamed as|rebranded as)\s*", "", part, flags=re.IGNORECASE).strip()
        if part and part.lower() != fallback_name.lower():
            cleaned.append(part)
    return cleaned


def decide_style(comments: str) -> str:
    text = comments.lower()
    if "1h" in text or "half" in text:
        return "half-year"
    if "q" in text:
        return "quarter"
    return "full-year"


def section_from_sheet2(record: Sheet2Record, mapped_lines: list[PortfolioLine]) -> str:
    if record.type_name.lower() == "fund":
        return "funds"

    if record.value_2025 == 0 and (record.invested or 0) > 0:
        return "writeoffs"

    sections = {line.section for line in mapped_lines}
    if "writeoffs" in sections:
        return "writeoffs"
    if "loans" in sections and "companies" not in sections:
        return "loans"
    return "companies"


def stable_asset_id(record: Sheet2Record, cleaned_name: str) -> str:
    norm = normalize_name(record.name)
    if norm in STABLE_ID_BY_SHEET2_NORM:
        return STABLE_ID_BY_SHEET2_NORM[norm]
    return slugify(cleaned_name)




def clamp_300(text: str) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= 300:
        return text
    return text[:297].rstrip() + "..."


def build_company_snapshot_300(name: str, description: str, latest_market_info_external: str, trend: str, diff: float | None) -> str:
    base = f"{name}: {description}."
    trend_text = ""
    if diff is not None:
        if trend == "growth":
            trend_text = f" Latest valuation trend is positive vs 1Q 2023 ({diff:,.0f} USD)."
        elif trend == "decline":
            trend_text = f" Latest valuation trend is negative vs 1Q 2023 ({diff:,.0f} USD)."
        else:
            trend_text = " Valuation is broadly stable versus 1Q 2023."
    market = f" External market view: {latest_market_info_external}" if latest_market_info_external else ""
    return clamp_300(base + trend_text + market)

def decline_reason(asset: dict[str, Any], diff: float | None, comments: str) -> str:
    if comments:
        return comments
    if diff is None:
        return "No differential data found in current sources."
    if diff >= 0:
        return "Not a decline."
    if asset.get("market_value_usd") == 0 and (asset.get("original_investment_usd") or 0) > 0:
        return "Marked to zero value in Dec 2025 valuation."
    if diff <= -5_000_000:
        return "Material valuation contraction versus 1Q 2023; needs source-backed root-cause writeup."
    return "Valuation is below 1Q 2023 baseline; clarification needed on drivers."


def build_dataset(xlsx_path: Path, pptx_path: Path, output_path: Path, overrides_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet2_records = build_sheet2_records(wb)
    portfolio_lines = build_portfolio_lines(wb)
    pptx_snippets = parse_pptx_text(pptx_path)
    overrides = load_overrides(overrides_path)

    sheet2_by_key: dict[str, Sheet2Record] = {normalize_name(rec.name): rec for rec in sheet2_records}
    lines_by_sheet2: dict[str, list[PortfolioLine]] = {k: [] for k in sheet2_by_key}

    for line in portfolio_lines:
        key = canonical_match_key(line.name)
        if key in lines_by_sheet2:
            lines_by_sheet2[key].append(line)

    assets: list[dict[str, Any]] = []

    for record in sheet2_records:
        key = normalize_name(record.name)
        mapped_lines = lines_by_sheet2.get(key, [])

        display_name = clean_company_name(record.name)
        asset_id = stable_asset_id(record, display_name)

        detailed_lines = [
            line
            for line in mapped_lines
            if line.nominal_holding not in (None, "", "n/a", "N/A")
        ]
        summary_lines = [line for line in mapped_lines if line not in detailed_lines]

        selected_lines = detailed_lines[:] if detailed_lines else summary_lines[:]

        if record.invested is not None and (detailed_lines or summary_lines):
            candidate_sets: list[list[PortfolioLine]] = []
            if detailed_lines:
                candidate_sets.append(detailed_lines)
            if summary_lines:
                best_summary = min(summary_lines, key=lambda line: abs((line.book_cost or 0) - record.invested))
                candidate_sets.append([best_summary])
            if detailed_lines and summary_lines:
                candidate_sets.append([best_summary, *detailed_lines])

            def score(lineset: list[PortfolioLine]) -> tuple[float, int]:
                total = sum((line.book_cost or 0) for line in lineset)
                # Minimize reconciliation gap, then prefer richer term-level breakdown.
                return (abs(total - record.invested), -len(lineset))

            selected_lines = min(candidate_sets, key=score)

        investments: list[dict[str, Any]] = []
        for line in selected_lines:
            meta = infer_instrument(line.name, line.description)
            investments.append(
                {
                    "security_name": line.name,
                    "instrument_type": meta["instrument_type"],
                    "series": meta["series"],
                    "share_class": meta["share_class"],
                    "nominal_holding": line.nominal_holding,
                    "average_cost": line.average_cost,
                    "book_cost_usd": line.book_cost,
                    "market_value_usd": line.market_value,
                    "pnl_usd": line.pnl,
                }
            )
        investments.sort(key=lambda x: x.get("book_cost_usd") or 0, reverse=True)

        tranche_book = sum((x.get("book_cost_usd") or 0) for x in investments)
        tranche_value = sum((x.get("market_value_usd") or 0) for x in investments)

        timeline: list[dict[str, Any]] = []
        if record.value_2023 is not None:
            timeline.append(
                {
                    "date": "2023-03-31",
                    "label": "1Q 2023 baseline",
                    "reporting_style": "quarter",
                    "value_usd": record.value_2023,
                    "source": "MVOF Master with Dec2025 valuations.xlsx / Sheet2",
                }
            )
        if record.value_2025 is not None:
            timeline.append(
                {
                    "date": "2025-12-31",
                    "label": "FY 2025 valuation",
                    "reporting_style": decide_style(record.comments),
                    "value_usd": record.value_2025,
                    "source": "MVOF Master with Dec2025 valuations.xlsx / Sheet2",
                }
            )

        timeline.append(
            {
                "date": "2026-04-24",
                "label": "2026 assessment reference",
                "event_type": "analysis",
                "reporting_style": "full-year",
                "value_usd": record.value_2025,
                "source": "MVOF 2026 Update.pptx",
            }
        )

        lineage: set[str] = set()
        for line in mapped_lines:
            for alias in parse_name_chain(line.previous_name, line.name):
                lineage.add(alias)

        for idx, alias in enumerate(sorted(lineage)):
            timeline.append(
                {
                    "date": f"2024-01-{min(idx + 1, 28):02d}",
                    "label": f"Name lineage includes: {alias}",
                    "event_type": "rename",
                    "reporting_style": "full-year",
                    "source": "MVOF Master with Dec2025 valuations.xlsx / Previous name",
                }
            )

        override = (overrides.get("assets", {}) or {}).get(asset_id, {})
        for event in (override.get("events", []) if isinstance(override, dict) else []):
            timeline.append(
                {
                    "date": event.get("date", "2026-01-01"),
                    "label": event.get("label", "Manual timeline event"),
                    "event_type": event.get("event_type", "note"),
                    "reporting_style": event.get("reporting_style", "full-year"),
                    "source": "canonical_overrides.json",
                }
            )

        timeline.sort(key=lambda item: item["date"], reverse=True)

        aliases = {record.name, display_name}
        for line in mapped_lines:
            aliases.add(line.name)
            if line.previous_name:
                aliases.update(parse_name_chain(line.previous_name, line.name))
        if isinstance(override, dict):
            aliases.update(override.get("aliases", []) or [])

        mention_probe = [display_name, record.name, *(aliases or [])]
        mention_count = 0
        for snippet in pptx_snippets:
            low = snippet.lower()
            if any(alias and alias.lower() in low for alias in mention_probe):
                mention_count += 1

        investment_types = sorted({x["instrument_type"] for x in investments if x.get("instrument_type")})
        structure_note = ""
        if investments:
            structure_note = f" Investment structure: {len(investments)} tranches ({', '.join(investment_types)})."

        section = section_from_sheet2(record, mapped_lines)
        market_value = record.value_2025 if record.value_2025 is not None else tranche_value
        original_investment = record.invested if record.invested is not None else tranche_book
        pnl = (market_value - original_investment) if (market_value is not None and original_investment is not None) else None

        latest_market_info_external = EXTERNAL_MARKET_INFO_BY_ID.get(asset_id, "") if section == "companies" else ""
        if latest_market_info_external:
            timeline.append(
                {
                    "date": "2026-03-31",
                    "label": "Latest market info: externally sourced information",
                    "event_type": "external_market_info",
                    "reporting_style": "full-year",
                    "source": "/Users/danielgusev/Desktop/MVOF Report.md",
                    "summary": latest_market_info_external,
                }
            )
            timeline.sort(key=lambda item: item["date"], reverse=True)

        spv_note = f" ({record.comments})" if record.comments else ""
        desc_text = (record.description + spv_note + structure_note).strip()
        trend_value = "growth" if (record.diff or 0) > 0 else "decline" if (record.diff or 0) < 0 else "stable"
        company_snapshot_sheet2 = record.company_snapshot.strip() if section == "companies" else ""
        company_snapshot_300 = build_company_snapshot_300(display_name, desc_text, latest_market_info_external, trend_value, record.diff) if section == "companies" else ""

        asset = {
            "id": asset_id,
            "name": display_name,
            "canonical_name": override.get("canonical_name") if isinstance(override, dict) and override.get("canonical_name") else display_name,
            "underlying_asset": override.get("underlying_asset") if isinstance(override, dict) and override.get("underlying_asset") else display_name,
            "aliases": sorted({a.strip() for a in aliases if a and str(a).strip()}),
            "section": section,
            "geography": mapped_lines[0].geography if mapped_lines and mapped_lines[0].geography else "",
            "description": desc_text,
            "company_snapshot_sheet2": company_snapshot_sheet2,
            "company_snapshot_300": company_snapshot_300,
            "previous_name": "; ".join(sorted(lineage)) if lineage else "",
            "original_investment_usd": original_investment,
            "value_2023_usd": record.value_2023,
            "book_cost_usd": tranche_book if tranche_book else original_investment,
            "market_value_usd": market_value,
            "pnl_usd": pnl,
            "diff_2025_vs_2023_usd": record.diff,
            "trend": trend_value,
            "major_slider": (record.diff or 0) <= -500000,
            "value_grower": (record.diff or 0) > 100000,
            "clarification_status": "Clarification needed" if record.diff is not None and abs(record.diff) <= 1.0 else "OK",
            "resolved": False,
            "decline_reason": "",
            "reporting_styles_available": sorted({entry["reporting_style"] for entry in timeline}),
            "timeline": timeline,
            "notes": [n for n in [record.comments, record.extra_comments] + [line.notes for line in mapped_lines if line.notes] if n],
            "investments": investments,
            "sheet2_reconciliation": {
                "sheet2_invested_usd": record.invested,
                "portfolio_tranche_book_sum_usd": tranche_book,
                "difference_usd": (record.invested - tranche_book) if (record.invested is not None and tranche_book) else None,
            },
            "latest_market_info_external": latest_market_info_external,
            "source_mentions": {
                "pptx_mentions": mention_count,
                "has_2026_analysis": mention_count > 0,
                "information_gap": mention_count == 0 and not investments and not record.comments,
            },
        }
        asset["decline_reason"] = decline_reason(asset, record.diff, record.comments)
        assets.append(asset)

    result = {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "sources": {
            "xlsx": str(xlsx_path),
            "pptx": str(pptx_path),
        },
        "password_hint": "Use scripts/set_password.py to configure hashed auth in data/auth.json",
        "assets": assets,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Build normalized MVOF dashboard dataset")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_PATHS["xlsx"])
    parser.add_argument("--pptx", type=Path, default=DEFAULT_PATHS["pptx"])
    parser.add_argument("--overrides", type=Path, default=DEFAULT_PATHS["overrides"])
    parser.add_argument("--output", type=Path, default=DEFAULT_PATHS["output"])
    args = parser.parse_args()

    data = build_dataset(args.xlsx, args.pptx, args.output, args.overrides)
    print(f"Wrote {len(data['assets'])} assets to {args.output}")


if __name__ == "__main__":
    main()
